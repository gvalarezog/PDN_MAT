from django.http import HttpResponse
from openpyxl import Workbook
from django.forms import modelform_factory
from django.shortcuts import render, get_object_or_404, redirect

# Create your views here.
from persona.forms import PersonaFormulario
from persona.models import Persona


def detalle_persona(request, id):
    # persona = Persona.objects.get(pk=id)
    persona = get_object_or_404(Persona, pk=id)
    mensaje = {'persona':persona}
    return render(request, 'detalle.html', mensaje)

#PersonaFormulario = modelform_factory(Persona, exclude=[])
def nueva_persona(request):
    if request.POST:
        formPersona = PersonaFormulario(request.POST)
        if formPersona.is_valid():
            formPersona.save()
            return redirect('inicio')
    else:
        formPersona = PersonaFormulario()
    return render(request, 'nueva.html', {'formPersona': formPersona})


def modificar_persona(request, id):
    persona = get_object_or_404(Persona, pk=id)
    if request.method == 'POST':
        formPersona = PersonaFormulario(request.POST, instance=persona)
        if formPersona.is_valid():
            formPersona.save()
            return redirect('inicio')
    else:
        formPersona = PersonaFormulario(instance=persona)
    mensaje = {'formPersona': formPersona}
    return render(request, 'modificar.html', context=mensaje)


def eliminar_persona(request, id):
    persona = get_object_or_404(Persona, pk=id)
    if persona:
        persona.delete()
    return redirect('inicio')

def reporte_personas(request):
    # Obtenemos todas las personas de nuestra base de datos
    personas = Persona.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE PERSONAS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    ws.title = 'Reporte'
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'ID'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO'
    ws['E3'] = 'EMAIL'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for persona in personas:
        ws.cell(row=cont, column=2).value = persona.id
        ws.cell(row=cont, column=3).value = persona.nombre
        ws.cell(row=cont, column=4).value = persona.apellido
        ws.cell(row=cont, column=5).value = persona.email
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReportePersonas.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
